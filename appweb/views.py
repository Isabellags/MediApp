from django.shortcuts import render, redirect, get_object_or_404
from .models import Profesional
from .forms import ContactoForm, ProfesionalForm
from django.contrib.auth.models import User, Group
from django.contrib.auth import authenticate, login
from django.http import HttpResponse
from openpyxl import Workbook

# Create your views here.
def home(request):

    return render(request, "home.html")

def contacto(request):

    data = {
        'miForm': ContactoForm
    }

    if request.method == "POST":
        formulario = ContactoForm(data = request.POST)

        if formulario.is_valid():
            formulario.save()
            data["mensaje"] = "Contacto guardado!!!"
        else:
            data['mensaje'] = "Hubo un problema"
            data['miForm'] = formulario

    return render(request, "contacto.html", data)





def profesionales(request):

    profs = Profesional.objects.all()

    #profesionales = Profesional.objects.raw("select * from appweb_profesional where especialista = true")


    data = {
        "mis_profesionales": profs,
        "saludo": "Holaaaa"
    }


    return render(request, "profesionales.html", data)


def agregar_profesional(request):


    data = {
        "form": ProfesionalForm
    }

    if request.method == "POST":

        formulario = ProfesionalForm(data=request.POST, files=request.FILES)

        if formulario.is_valid():
            formulario.save()
            data["mensaje"] = "Registro creado correctamente!"
        else:
            data["mensaje"] = "hubo un error!!!!"
            data["form"] = formulario
            


    return render(request, "mantenedor/profesional/agregar.html", data)


def listar_profesional(request):

    mis_profesionales = Profesional.objects.all()

    data = {
        "profesionales" : mis_profesionales
    }


    return render(request, "mantenedor/profesional/listar.html", data)


def modificar_profesional(request, rut):

    profesional = get_object_or_404(Profesional, rut=rut)

    data = {
        'form': ProfesionalForm(instance=profesional)
    }

    if request.method == 'POST':
        formulario = ProfesionalForm(data=request.POST, files=request.FILES, instance=profesional)

        if formulario.is_valid():
            formulario.save()
            return redirect(to="listar_profesional")
        else:
            data["mensaje"] = "No se puede guardar"
            data["form"] = formulario


    return render(request, 'mantenedor/profesional/modificar.html', data)


def eliminar_profesional(request, rut):

    profesional = get_object_or_404(Profesional, rut=rut)

    profesional.delete()

    return redirect(to="listar_profesional")

def login_usuario(request):
    print("USUARIO: "+ request.user.username)

    print("Grupos ",  request.user.groups.all())

    if request.user.groups.filter(name='profesional'):
        print("Es un profesional")



    return redirect(to='home')

def registro_profesional(request):

    data = {
        "mensaje" : ""
    }

    if request.POST:
        nombre = request.POST.get("nombre")
        apellido = request.POST.get("apellido")
        correo = request.POST.get("correo")
        password1 = request.POST.get("password1")
        password2 = request.POST.get("password2")
        if password1 != password2:
            data["mensaje"] = "Las contraseñas deben ser identicas"
        else:
            usu = User()
            usu.set_password(password1)
            usu.email = correo
            usu.username = nombre
            usu.first_name = nombre
            usu.last_name = apellido
            grupo = Group.objects.get(name='profesional')
            try:
                usu.save()
                usu.groups.add(grupo)
                data["mensaje"] = "Usuario creado"

                user = authenticate(username=usu.username, password=password1)
                login(request, user)
                return redirect(to='home')

            except:
                data["mensaje"] = "Hubo un error"

        
    return render(request, "registration/registro.html", data)



def reporte_profesionales_por_fecha(request):
    # Ejemplo de filtro por mes de junio
    profesionales = Profesional.objects.filter(fecha_nacimiento__month=6)

    # Renderizar el template del reporte con los datos filtrados
    return render(request, 'reporte_profesionales.html', {'profesionales': profesionales})



def descargar_excel_profesionales_por_fecha(request):
    # Ejemplo de filtro por mes de junio
    profesionales = Profesional.objects.filter(fecha_nacimiento__month=6)

    # Crear el libro de Excel y la hoja
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Profesionales por Fecha"

    # Headers
    ws['A1'] = "RUT"
    ws['B1'] = "Nombre"
    ws['C1'] = "Apellido"
    ws['D1'] = "Edad"
    ws['E1'] = "Especialista"
    ws['F1'] = "Cargo"
    ws['G1'] = "Fecha de Nacimiento"

    # Datos de profesionales filtrados
    row = 2
    for profesional in profesionales:
        ws[f'A{row}'] = profesional.rut
        ws[f'B{row}'] = profesional.nombre
        ws[f'C{row}'] = profesional.apellido
        ws[f'D{row}'] = profesional.edad
        ws[f'E{row}'] = "Sí" if profesional.especialista else "No"
        ws[f'F{row}'] = str(profesional.cargo)
        ws[f'G{row}'] = str(profesional.fecha_nacimiento)
        row += 1

    # Guardar el libro de Excel
    response = HttpResponse(content_type='application/vnd.openpyxl.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="reporte_profesionales_junio.xlsx"'
    wb.save(response)
    return response